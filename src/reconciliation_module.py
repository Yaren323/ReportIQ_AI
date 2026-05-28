from __future__ import annotations

import re
import hashlib
from datetime import datetime
from io import BytesIO

import pandas as pd
import streamlit as st
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

UNIVERA_SHEET = "Univera Envanter"
TF_SHEET_UNIVIS = "Univera&Univis Sunucu Listesi"
TF_SHEET_ODM = "Univera ODM Sunucu Listesi"
TF_SOURCE_COL = "Türk Finansal Kaynak Sheet"


def _norm_text(value: str) -> str:
    return re.sub(r"\s+", " ", str(value).strip().lower())


def normalize_server_key(value):
    if pd.isna(value):
        return None
    value = str(value).strip()
    value = value.replace("\u00a0", " ")
    value = " ".join(value.split())
    if value == "":
        return None
    lower = value.lower().strip()
    invalid_values = [
        "nan",
        "none",
        "null",
        "toplam",
        "total",
        "genel toplam",
        "ara toplam",
        "hostname",
        "vmname",
        "sunucu adı",
        "server name",
        "aciklama satiri",
        "açıklama satırı",
        "baslik satiri",
        "başlık satırı",
    ]
    if lower in invalid_values:
        return None
    if re.fullmatch(r"\d+", lower):
        return None
    if value.endswith(".0"):
        value = value[:-2]
    return value.upper()


def soft_normalize_server_key(value):
    base = normalize_server_key(value)
    if base is None:
        return None
    soft = re.sub(r"[^A-Z0-9]", "", base)
    return soft if soft else None


def _parse_number(value) -> float | None:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None

    text = str(value).strip().lower().replace(" ", "")
    if not text:
        return None

    unit = ""
    if "tb" in text:
        unit = "tb"
    elif "gb" in text:
        unit = "gb"
    elif "mb" in text:
        unit = "mb"

    num_text = re.sub(r"[^0-9,\.-]", "", text)
    if not num_text:
        return None

    if "," in num_text and "." in num_text:
        if num_text.rfind(",") > num_text.rfind("."):
            num_text = num_text.replace(".", "").replace(",", ".")
        else:
            num_text = num_text.replace(",", "")
    elif "," in num_text:
        num_text = num_text.replace(",", ".") if num_text.count(",") == 1 else num_text.replace(",", "")

    try:
        num = float(num_text)
    except ValueError:
        return None

    if unit == "tb":
        return num * 1024
    if unit == "mb":
        return num / 1024
    return num


def _diff_status(u_val: float | None, t_val: float | None) -> tuple[str, float | None]:
    if u_val is None or t_val is None:
        return "Kontrol Edilmeli", None
    if u_val == t_val:
        return "Eşit", 0.0
    diff = abs(u_val - t_val)
    return (f"{diff:g} eksik", diff) if u_val < t_val else (f"{diff:g} fazla", diff)


def _detect_col(df: pd.DataFrame, candidates: list[str]) -> str | None:
    col_map = {_norm_text(c): c for c in df.columns}
    for c in candidates:
        key = _norm_text(c)
        if key in col_map:
            return col_map[key]
    for col in df.columns:
        ncol = _norm_text(col)
        if any(_norm_text(c) in ncol for c in candidates):
            return col
    return None


def _autosize_and_style(ws, header_color="1F4E78", diff_columns: list[str] | None = None):
    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            max_len = max(max_len, len(str(val)) if val is not None else 0)
        ws.column_dimensions[col_letter].width = min(max(14, max_len + 2), 56)

    if not diff_columns:
        return

    headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
    targets = [headers.index(name) + 1 for name in diff_columns if name in headers]
    diff_fill = PatternFill(start_color="FDE2E1", end_color="FDE2E1", fill_type="solid")

    for r in range(2, ws.max_row + 1):
        for c in targets:
            txt = str(ws.cell(row=r, column=c).value or "")
            if txt not in {"Eşit", "Tam Eşleşti"}:
                ws.cell(row=r, column=c).fill = diff_fill


def _build_tf_combined(univis_df: pd.DataFrame, odm_df: pd.DataFrame) -> pd.DataFrame:
    part_univis = univis_df.copy()
    part_univis[TF_SOURCE_COL] = TF_SHEET_UNIVIS
    part_univis["_excel_row_no"] = part_univis.index + 2
    part_odm = odm_df.copy()
    part_odm[TF_SOURCE_COL] = TF_SHEET_ODM
    part_odm["_excel_row_no"] = part_odm.index + 2
    return pd.concat([part_univis, part_odm], ignore_index=True)


def _run_reconciliation(
    u_df: pd.DataFrame,
    tf_df: pd.DataFrame,
    univis_raw_rows: int,
    odm_raw_rows: int,
    univera_file_name: str,
    tf_file_name: str,
):
    u_host_col = next((c for c in u_df.columns if _norm_text(c) == "hostname"), None)
    t_host_col = _detect_col(tf_df, ["VMName", "VM Name", "Hostname"])

    u_ram_col = _detect_col(u_df, ["RAM", "Memory", "Bellek"])
    t_ram_col = _detect_col(tf_df, ["RAM", "Memory", "Bellek"])
    u_cpu_col = _detect_col(u_df, ["CPU", "Core", "vCPU", "İşlemci", "Islemci"])
    t_cpu_col = _detect_col(tf_df, ["CPU", "Core", "vCPU", "İşlemci", "Islemci"])
    u_disk_col = _detect_col(u_df, ["Disk", "Storage", "HDD", "SSD", "Kapasite"])
    t_disk_col = _detect_col(tf_df, ["Disk", "Storage", "HDD", "SSD", "Kapasite"])

    if not u_host_col:
        raise ValueError("Univera Envanter sheet'inde tam adı 'Hostname' olan kolon bulunamadı.")
    if not t_host_col:
        raise ValueError("Türk Finansal sheetlerinde VMName/Hostname kolonu bulunamadı.")

    u_raw_row_count = len(u_df)
    t_raw_row_count = len(tf_df)
    u = u_df.copy()
    t = tf_df.copy()
    u["_excel_row_no"] = u.index + 2

    u["_match_key"] = u[u_host_col].apply(normalize_server_key)
    t["_match_key"] = t[t_host_col].apply(normalize_server_key)
    u["_soft_key"] = u[u_host_col].apply(soft_normalize_server_key)
    t["_soft_key"] = t[t_host_col].apply(soft_normalize_server_key)

    u = u[u["_match_key"].notna()].copy()
    t = t[t["_match_key"].notna()].copy()

    u_unique = u.drop_duplicates(subset=["_match_key"]).copy()
    t_unique = t.drop_duplicates(subset=["_match_key"]).copy()
    u_duplicate_count = int(u["_match_key"].duplicated().sum())
    t_duplicate_count = int(t["_match_key"].duplicated().sum())

    u_map = u_unique.set_index("_match_key")
    t_map = t_unique.set_index("_match_key")

    u_keys = set(u_unique["_match_key"].astype(str))
    t_keys = set(t_unique["_match_key"].astype(str))

    primary_matches = sorted(u_keys & t_keys)
    u_only = set(u_keys - t_keys)
    t_only = set(t_keys - u_keys)

    u_soft_map = u_unique[u_unique["_match_key"].isin(u_only)].set_index("_match_key")["_soft_key"].to_dict()
    t_soft_map = t_unique[t_unique["_match_key"].isin(t_only)].set_index("_match_key")["_soft_key"].to_dict()

    secondary_pairs = []
    t_soft_reverse = {}
    for t_key, soft in t_soft_map.items():
        if soft:
            t_soft_reverse.setdefault(soft, []).append(t_key)

    used_t_secondary = set()
    for u_key, soft in u_soft_map.items():
        if not soft:
            continue
        candidates = t_soft_reverse.get(soft, [])
        candidate = next((c for c in candidates if c not in used_t_secondary), None)
        if candidate:
            secondary_pairs.append((u_key, candidate))
            used_t_secondary.add(candidate)

    for u_key, t_key in secondary_pairs:
        u_only.discard(u_key)
        t_only.discard(t_key)

    match_pairs = [(k, k, "Ana eşleşme") for k in primary_matches]
    match_pairs.extend([(u_key, t_key, "İkinci kontrol ile eşleşti") for u_key, t_key in secondary_pairs])

    matched_rows = []
    diff_rows = []

    for u_key, t_key, match_type in match_pairs:
        u_row = u_map.loc[u_key]
        t_row = t_map.loc[t_key]

        u_ram = _parse_number(u_row[u_ram_col]) if u_ram_col else None
        t_ram = _parse_number(t_row[t_ram_col]) if t_ram_col else None
        ram_status, ram_diff = _diff_status(u_ram, t_ram)

        u_cpu = _parse_number(u_row[u_cpu_col]) if u_cpu_col else None
        t_cpu = _parse_number(t_row[t_cpu_col]) if t_cpu_col else None
        cpu_status, cpu_diff = _diff_status(u_cpu, t_cpu)

        u_disk = _parse_number(u_row[u_disk_col]) if u_disk_col else None
        t_disk = _parse_number(t_row[t_disk_col]) if t_disk_col else None
        disk_status, disk_diff = _diff_status(u_disk, t_disk)

        if all(s == "Eşit" for s in [ram_status, cpu_status, disk_status]):
            overall = "Tam Eşleşti"
        elif any(s == "Kontrol Edilmeli" for s in [ram_status, cpu_status, disk_status]):
            overall = "Kontrol Edilmeli"
        else:
            overall = "Fark Var"

        server_name = u_row[u_host_col] if pd.notna(u_row[u_host_col]) else t_row[t_host_col]

        matched_rows.append(
            {
                "Sunucu Adı": normalize_server_key(server_name),
                TF_SOURCE_COL: t_row.get(TF_SOURCE_COL, ""),
                "Eşleşme Tipi": match_type,
                "Univera RAM": u_ram,
                "Türk Finansal RAM": t_ram,
                "RAM Durumu": ram_status,
                "Univera CPU": u_cpu,
                "Türk Finansal CPU": t_cpu,
                "CPU Durumu": cpu_status,
                "Univera Disk": u_disk,
                "Türk Finansal Disk": t_disk,
                "Disk Durumu": disk_status,
                "Genel Durum": overall,
            }
        )

        for field, u_val, t_val, status, diff in [
            ("RAM", u_ram, t_ram, ram_status, ram_diff),
            ("CPU", u_cpu, t_cpu, cpu_status, cpu_diff),
            ("Disk", u_disk, t_disk, disk_status, disk_diff),
        ]:
            if status != "Eşit":
                diff_rows.append(
                    {
                        "Sunucu Adı": normalize_server_key(server_name),
                        TF_SOURCE_COL: t_row.get(TF_SOURCE_COL, ""),
                        "Alan": field,
                        "Univera Değeri": u_val,
                        "Türk Finansal Değeri": t_val,
                        "Fark Durumu": status,
                        "Fark Miktarı": diff,
                    }
                )

    matched_df = pd.DataFrame(matched_rows)
    if not matched_df.empty:
        matched_df = matched_df.sort_values(by=["Sunucu Adı", TF_SOURCE_COL], na_position="last").reset_index(drop=True)

    # Güvenlik: only listeleri yalnızca gerçekten ilgili tarafta bulunan key'lerden üretilir.
    valid_u_only = sorted([k for k in u_only if k in u_keys])
    u_only_df = pd.DataFrame(
        {
            "Sunucu Adı": valid_u_only,
            "Univera Sheet": UNIVERA_SHEET,
            "Durum": "Univera 2026 listesinde var, Türk Finansal listesinde yok",
        }
    )

    valid_t_only = set([k for k in t_only if k in t_keys])
    t_only_source = t_unique[t_unique["_match_key"].isin(valid_t_only)][["_match_key", TF_SOURCE_COL]].drop_duplicates()
    t_only_df = pd.DataFrame(
        {
            "Sunucu Adı": t_only_source["_match_key"].astype(str) if not t_only_source.empty else [],
            TF_SOURCE_COL: t_only_source[TF_SOURCE_COL] if not t_only_source.empty else [],
            "Durum": "Türk Finansal listesinde var, Univera 2026 listesinde yok",
        }
    )

    diff_df = pd.DataFrame(diff_rows)
    if not diff_df.empty:
        diff_df = diff_df.sort_values(by=["Sunucu Adı", "Alan"], na_position="last").reset_index(drop=True)

    total_u = int(u_unique["_match_key"].nunique())
    total_t = int(t_unique["_match_key"].nunique())
    matched = len(matched_df)
    u_only_count = len(u_only_df)
    t_only_count = len(t_only_df)
    fark_count = len(matched_df[matched_df["Genel Durum"] == "Fark Var"]) if not matched_df.empty else 0
    tam_eslesen = len(matched_df[matched_df["Genel Durum"] == "Tam Eşleşti"]) if not matched_df.empty else 0
    rate = (tam_eslesen / matched * 100) if matched > 0 else 0.0

    kpi = {
        "Univera 2026 toplam sunucu sayısı": total_u,
        "Türk Finansal birleşik toplam sunucu sayısı": total_t,
        "Eşleşen sunucu sayısı": matched,
        "Univera’da olup Türk Finansalda olmayan sunucu sayısı": u_only_count,
        "Türk Finansalda olup Univera’da olmayan sunucu sayısı": t_only_count,
        "Fark bulunan sunucu sayısı": fark_count,
        "Tam eşleşen sunucu sayısı": tam_eslesen,
        "Mutabakat oranı": f"{rate:.2f}%",
    }
    u_disk_total = (
        u_unique[u_disk_col].apply(_parse_number).dropna().sum()
        if u_disk_col and u_disk_col in u_unique.columns
        else 0.0
    )
    t_disk_total = (
        t_unique[t_disk_col].apply(_parse_number).dropna().sum()
        if t_disk_col and t_disk_col in t_unique.columns
        else 0.0
    )

    debug_summary = {
        "Univera ham satır sayısı": u_raw_row_count,
        "Univera geçerli tekil Hostname sayısı": total_u,
        "Türk Finansal Univis ham satır sayısı": univis_raw_rows,
        "Türk Finansal ODM ham satır sayısı": odm_raw_rows,
        "Türk Finansal birleşik ham satır sayısı": t_raw_row_count,
        "Türk Finansal geçerli tekil VMName sayısı": total_t,
        "Univera duplicate Hostname sayısı": u_duplicate_count,
        "Türk Finansal duplicate VMName sayısı": t_duplicate_count,
        "Univera disk toplamı": round(float(u_disk_total), 2),
        "Türk Finansal disk toplamı": round(float(t_disk_total), 2),
        "İlk 10 Univera Hostname": ", ".join(list(u_unique["_match_key"].astype(str).head(10))),
        "İlk 10 Türk Finansal VMName": ", ".join(list(t_unique["_match_key"].astype(str).head(10))),
    }
    u_debug_df = (
        u_unique[[u_host_col, "_match_key", "_excel_row_no"]]
        .rename(
            columns={
                u_host_col: "Ham Hostname",
                "_match_key": "Normalize Key",
                "_excel_row_no": "Orijinal Excel Satır No",
            }
        )
        .copy()
    )
    u_debug_df.insert(0, "Sıra", range(1, len(u_debug_df) + 1))
    u_debug_df["Geldiği Sheet"] = UNIVERA_SHEET
    u_debug_df["Geldiği Dosya"] = univera_file_name
    u_debug_df = u_debug_df[
        ["Sıra", "Ham Hostname", "Normalize Key", "Geldiği Sheet", "Geldiği Dosya", "Orijinal Excel Satır No"]
    ]

    t_debug_df = (
        t_unique[[t_host_col, "_match_key", TF_SOURCE_COL, "_excel_row_no"]]
        .rename(
            columns={
                t_host_col: "Ham VMName",
                "_match_key": "Normalize Key",
                TF_SOURCE_COL: "Geldiği Sheet",
                "_excel_row_no": "Orijinal Excel Satır No",
            }
        )
        .copy()
    )
    t_debug_df.insert(0, "Sıra", range(1, len(t_debug_df) + 1))
    t_debug_df["Geldiği Dosya"] = tf_file_name
    t_debug_df = t_debug_df[
        ["Sıra", "Ham VMName", "Normalize Key", "Geldiği Sheet", "Geldiği Dosya", "Orijinal Excel Satır No"]
    ]

    u_valid_count = int(u["_match_key"].notna().sum())
    univis_valid_count = int(t[t[TF_SOURCE_COL] == TF_SHEET_UNIVIS]["_match_key"].notna().sum())
    odm_valid_count = int(t[t[TF_SOURCE_COL] == TF_SHEET_ODM]["_match_key"].notna().sum())
    sheet_usage_df = pd.DataFrame(
        [
            {
                "Dosya": univera_file_name,
                "Sheet": UNIVERA_SHEET,
                "Ham Satır Sayısı": u_raw_row_count,
                "Geçerli Sunucu Sayısı": u_valid_count,
            },
            {
                "Dosya": tf_file_name,
                "Sheet": TF_SHEET_UNIVIS,
                "Ham Satır Sayısı": univis_raw_rows,
                "Geçerli Sunucu Sayısı": univis_valid_count,
            },
            {
                "Dosya": tf_file_name,
                "Sheet": TF_SHEET_ODM,
                "Ham Satır Sayısı": odm_raw_rows,
                "Geçerli Sunucu Sayısı": odm_valid_count,
            },
        ]
    )

    debug_tables = {
        "univera_servers": u_debug_df,
        "tf_servers": t_debug_df,
        "sheet_usage": sheet_usage_df,
    }

    return kpi, matched_df, u_only_df, t_only_df, diff_df, debug_summary, debug_tables


def _generate_excel(kpi: dict, matched_df: pd.DataFrame, u_only_df: pd.DataFrame, t_only_df: pd.DataFrame, diff_df: pd.DataFrame) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary_df = pd.DataFrame(
            [
                {"Metrik": "Rapor Başlığı", "Değer": "ReportIQ Mutabakat ve Kapasite Karşılaştırma Raporu"},
                {"Metrik": "Rapor Tarihi", "Değer": datetime.now().strftime("%Y-%m-%d %H:%M")},
                *[{"Metrik": k, "Değer": v} for k, v in kpi.items()],
            ]
        )

        summary_df.to_excel(writer, sheet_name="Yönetici Özeti", index=False)
        matched_df.to_excel(writer, sheet_name="Eşleşen Sunucular", index=False)
        u_only_df.to_excel(writer, sheet_name="Univera Olup Türk Yok", index=False)
        t_only_df.to_excel(writer, sheet_name="Türk Olup Univera Yok", index=False)
        diff_df.to_excel(writer, sheet_name="Fark Analizi", index=False)

        _autosize_and_style(writer.book["Yönetici Özeti"], header_color="243447")
        _autosize_and_style(
            writer.book["Eşleşen Sunucular"],
            header_color="1F4E78",
            diff_columns=["RAM Durumu", "CPU Durumu", "Disk Durumu", "Genel Durum"],
        )
        _autosize_and_style(writer.book["Univera Olup Türk Yok"], header_color="1F4E78")
        _autosize_and_style(writer.book["Türk Olup Univera Yok"], header_color="1F4E78")
        _autosize_and_style(writer.book["Fark Analizi"], header_color="8B4513", diff_columns=["Fark Durumu"])

    return output.getvalue()


def render_reconciliation_module():
    st.title("AI Mutabakat ve Kapasite Karşılaştırma")
    st.caption("Veriler kaydedilmez. Yükleme sonrası analiz edilir ve rapor indirilebilir.")

    c1, c2 = st.columns(2)
    with c1:
        univera_file = st.file_uploader("Univera 2026 Excel", type=["xlsx", "xls"], key="recon_univera")
    with c2:
        tf_file = st.file_uploader("Türk Finansal Excel", type=["xlsx", "xls"], key="recon_finansal")

    if not (univera_file and tf_file):
        st.info("Analiz için iki Excel dosyasını yükleyin.")
        return

    # Dosya değiştiyse eski analiz sonuçlarını temizle.
    u_bytes = univera_file.getvalue()
    t_bytes = tf_file.getvalue()
    current_signature = hashlib.md5(u_bytes + b"||" + t_bytes).hexdigest()
    previous_signature = st.session_state.get("reconciliation_file_signature")
    if previous_signature != current_signature:
        for key in [
            "reconciliation_last_result",
            "reconciliation_file_signature",
        ]:
            st.session_state.pop(key, None)
        st.session_state["reconciliation_file_signature"] = current_signature

    if not st.button("Mutabakat Analizi Çalıştır", key="run_reconciliation"):
        last_result = st.session_state.get("reconciliation_last_result")
        if last_result is None:
            return
        kpi, matched_df, u_only_df, t_only_df, diff_df, debug_summary, debug_tables = last_result
    else:
        try:
            u_sheets = pd.read_excel(BytesIO(u_bytes), sheet_name=None)
            tf_sheets = pd.read_excel(BytesIO(t_bytes), sheet_name=None)
        except Exception as exc:
            st.error(f"Excel dosyası okunamadı: {exc}")
            return

        if UNIVERA_SHEET not in u_sheets:
            st.error(f"Univera dosyasında '{UNIVERA_SHEET}' sheet'i bulunamadı.")
            return
        if TF_SHEET_UNIVIS not in tf_sheets or TF_SHEET_ODM not in tf_sheets:
            st.error(
                f"Türk Finansal dosyasında gerekli sheetler bulunamadı: '{TF_SHEET_UNIVIS}' ve '{TF_SHEET_ODM}'"
            )
            return

        u_sheet_df = u_sheets[UNIVERA_SHEET].copy()
        univis_df = tf_sheets[TF_SHEET_UNIVIS].copy()
        odm_df = tf_sheets[TF_SHEET_ODM].copy()
        tf_combined = _build_tf_combined(univis_df, odm_df)

        try:
            kpi, matched_df, u_only_df, t_only_df, diff_df, debug_summary, debug_tables = _run_reconciliation(
                u_sheet_df,
                tf_combined,
                univis_raw_rows=len(univis_df),
                odm_raw_rows=len(odm_df),
                univera_file_name=univera_file.name,
                tf_file_name=tf_file.name,
            )
        except ValueError as exc:
            st.error(str(exc))
            return

        st.session_state["reconciliation_last_result"] = (
            kpi,
            matched_df,
            u_only_df,
            t_only_df,
            diff_df,
            debug_summary,
            debug_tables,
        )
    st.caption(f"Kullanılan Türk Finansal sheetler: {TF_SHEET_UNIVIS}, {TF_SHEET_ODM}")

    m1, m2, m3, m4 = st.columns(4)
    m5, m6, m7, m8 = st.columns(4)
    m1.metric("Univera Toplam", kpi["Univera 2026 toplam sunucu sayısı"])
    m2.metric("Türk Finansal Toplam", kpi["Türk Finansal birleşik toplam sunucu sayısı"])
    m3.metric("Eşleşen", kpi["Eşleşen sunucu sayısı"])
    m4.metric("Mutabakat Oranı", kpi["Mutabakat oranı"])
    m5.metric("Sadece Univera", kpi["Univera’da olup Türk Finansalda olmayan sunucu sayısı"])
    m6.metric("Sadece Türk Finansal", kpi["Türk Finansalda olup Univera’da olmayan sunucu sayısı"])
    m7.metric("Fark Bulunan", kpi["Fark bulunan sunucu sayısı"])
    m8.metric("Tam Eşleşen", kpi["Tam eşleşen sunucu sayısı"])

    st.markdown("### Eşleşen Sunucular")
    if matched_df.empty:
        st.info("Eşleşen sunucu bulunamadı.")
    else:
        def _hl(row):
            if row.get("Genel Durum") != "Tam Eşleşti":
                return ["background-color: #FDE2E1"] * len(row)
            return [""] * len(row)

        st.dataframe(matched_df.style.apply(_hl, axis=1), use_container_width=True, hide_index=True)

    st.markdown("### Univera’da Olup Türk Finansalda Olmayan Sunucular")
    st.dataframe(u_only_df, use_container_width=True, hide_index=True)

    st.markdown("### Türk Finansalda Olup Univera’da Olmayan Sunucular")
    st.dataframe(t_only_df, use_container_width=True, hide_index=True)

    st.markdown("### Fark Analizi")
    if diff_df.empty:
        st.success("Fark bulunan kayıt yok.")
    else:
        st.dataframe(diff_df, use_container_width=True, hide_index=True)

    st.markdown("### Veri Kontrol Özeti")
    debug_df = pd.DataFrame(
        [{"Metrik": metric, "Değer": value} for metric, value in debug_summary.items()]
    )
    st.dataframe(debug_df, use_container_width=True, hide_index=True)

    st.markdown("### DEBUG AKTİF")
    st.warning("DEBUG AKTİF - KPI bu tablodaki normalize key sayısından hesaplanıyor")

    st.markdown("#### DEBUG - Univera olarak sayılan sunucular")
    st.dataframe(debug_tables["univera_servers"], use_container_width=True, hide_index=True)

    st.markdown("#### DEBUG - Türk Finansal olarak sayılan sunucular")
    st.dataframe(debug_tables["tf_servers"], use_container_width=True, hide_index=True)

    st.markdown("#### DEBUG - Kullanılan Sheetler")
    st.dataframe(debug_tables["sheet_usage"], use_container_width=True, hide_index=True)

    excel_bytes = _generate_excel(kpi, matched_df, u_only_df, t_only_df, diff_df)
    st.download_button(
        "Raporu Excel İndir",
        data=excel_bytes,
        file_name="reportiq_mutabakat_raporu.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="download_simple_reconciliation_report",
    )
